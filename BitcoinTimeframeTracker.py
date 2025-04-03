import asyncio
import aiohttp
import datetime
import logging
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt

# Définition des couleurs ANSI
YELLOW = "\033[1;33m"  # Jaune (or)
RED = "\033[1;31m"     # Rouge
NORMAL = "\033[0m"     # Réinitialisation des couleurs

# Configuration de la journalisation
logging.basicConfig(filename='crypto_tracker.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Initialisation des variables globales
unesminutesdebut = 3
troisminutesdebut = 3
cinqsminutesdebut = 3
quinsesminutesdebut = 3
trenteminutesdebut = 3
uneheuredebut = 3
deuxheuresdebut = 3
quatreheuresdebut = 3
sixheuresdebut = 3
huitheuresdebut = 3
douzeheuresdebut = 3
unedjoudebut = 3
troisjousdebut = 3
une_semaine_debut = 3

lancements = {
    "lancement3minutes": 0,
    "lancement5minutes": 0,
    "lancement15minutes": 0,
    "lancement30minutes": 0,
    "lancement1h": 0,
    "lancement2h": 0,
    "lancement4h": 0,
    "lancement6h": 0,
    "lancement8h": 0,
    "lancement12h": 0,
    "lancement1jour": 0,
    "lancement3jour": 0,
    "lancement1semaine": 0,
    "total": 0
}

timeframes = {
    "1m": "1m",
    "3m": "3m",
    "5m": "5m",
    "15m": "15m",
    "30m": "30m",
    "1h": "1h",
    "2h": "2h",
    "4h": "4h",
    "6h": "6h",
    "8h": "8h",
    "12h": "12h",
    "1d": "1d",
    "3d": "3d",
    "1w": "1w",
}

wb = Workbook()
sheet = wb.active
alignment = Alignment(horizontal='center', vertical='center')

row = 1
debutlongueur = 1
longueur = 6

for i, (timeframe, value) in enumerate(timeframes.items()):
    sheet.merge_cells(start_row=row, start_column=debutlongueur, end_row=row, end_column=longueur)
    cell = sheet.cell(row=row, column=debutlongueur, value=f"Time Frame : {timeframe}")
    cell.alignment = alignment

    sheet.cell(row=row + 1, column=debutlongueur, value="timestamp").alignment = alignment
    sheet.cell(row=row + 1, column=debutlongueur + 1, value="Open").alignment = alignment
    sheet.cell(row=row + 1, column=debutlongueur + 2, value="High").alignment = alignment
    sheet.cell(row=row + 1, column=debutlongueur + 3, value="Low").alignment = alignment
    sheet.cell(row=row + 1, column=debutlongueur + 4, value="Close").alignment = alignment
    sheet.cell(row=row + 1, column=debutlongueur + 5, value="volume").alignment = alignment

    debutlongueur = longueur + 1
    longueur = longueur + 6

async def fetch_ohlcv(session, symbol, timeframe):
    url = f"https://api.binance.com/api/v3/klines?symbol={symbol}&interval={timeframe}"
    try:
        async with session.get(url) as response:
            data = await response.json()
            return data
    except Exception as e:
        logging.error(f"Erreur lors de la récupération des données OHLCV pour {timeframe}: {e}")
        return None

async def update_sheet(timeframe, row, col, save_csv):
    async with aiohttp.ClientSession() as session:
        ohlcv_data = await fetch_ohlcv(session, 'BTCUSDT', timeframe)
        if ohlcv_data is None:
            return

        last_candle = ohlcv_data[-1]

        timestamp = datetime.datetime.utcfromtimestamp(last_candle[0] / 1000).strftime('%Y-%m-%d %H:%M:%S')
        open_price = last_candle[1]
        high_price = last_candle[2]
        low_price = last_candle[3]
        close_price = last_candle[4]
        volume = last_candle[5]

        sheet.cell(row=row, column=col, value=timestamp).alignment = alignment
        sheet.cell(row=row, column=col + 1, value=open_price).alignment = alignment
        sheet.cell(row=row, column=col + 2, value=high_price).alignment = alignment
        sheet.cell(row=row, column=col + 3, value=low_price).alignment = alignment
        sheet.cell(row=row, column=col + 4, value=close_price).alignment = alignment
        sheet.cell(row=row, column=col + 5, value=volume).alignment = alignment

        # Sauvegarde des données dans un fichier CSV si demandé
        if save_csv:
            save_to_csv(timeframe, [timestamp, open_price, high_price, low_price, close_price, volume])

        # Mise à jour du graphique
        update_plot(timeframe, ohlcv_data)

def save_to_csv(timeframe, data):
    filename = "crypto_data.csv"
    try:
        file_exists = False
        try:
            with open(filename, mode='r'):
                file_exists = True
        except FileNotFoundError:
            pass

        with open(filename, mode='a', newline='') as file:
            writer = csv.writer(file)
            if not file_exists:
                writer.writerow(["timeframe", "timestamp", "open", "high", "low", "close", "volume"])
            writer.writerow([timeframe] + data)
    except Exception as e:
        logging.error(f"Erreur lors de l'écriture dans le fichier CSV {filename}: {e}")

def save_to_excel():
    try:
        df = pd.read_csv("crypto_data.csv")
        df.to_excel("crypto_data.xlsx", index=False)
    except Exception as e:
        logging.error(f"Erreur lors de la conversion en Excel: {e}")

def update_plot(timeframe, ohlcv_data):
    timestamps = [datetime.datetime.utcfromtimestamp(candle[0] / 1000) for candle in ohlcv_data]
    close_prices = [candle[4] for candle in ohlcv_data]

    plt.figure(figsize=(10, 5))
    plt.plot(timestamps, close_prices, label=f'Close Price ({timeframe})')
    plt.xlabel('Time')
    plt.ylabel('Close Price')
    plt.title(f'BTCUSDT Close Prices ({timeframe})')
    plt.legend()
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(f'btcusdt_{timeframe}.png')
    plt.close()

async def main(save_csv, save_excel):
    global unesminutesdebut, troisminutesdebut, cinqsminutesdebut, quinsesminutesdebut, trenteminutesdebut
    global uneheuredebut, deuxheuresdebut, quatreheuresdebut, sixheuresdebut, huitheuresdebut
    global douzeheuresdebut, unedjoudebut, troisjousdebut, une_semaine_debut

    while True:
        await update_sheet("1m", unesminutesdebut, 1, save_csv)
        unesminutesdebut += 1
        if save_excel:
            wb.save("DATA_crypto.xlsx")

        for key in lancements:
            lancements[key] += 1

        if lancements["lancement3minutes"] == 3:
            await update_sheet("3m", troisminutesdebut, 7, save_csv)
            troisminutesdebut += 1
            lancements["lancement3minutes"] = 0

        if lancements["lancement5minutes"] == 5:
            await update_sheet("5m", cinqsminutesdebut, 13, save_csv)
            cinqsminutesdebut += 1
            lancements["lancement5minutes"] = 0

        if lancements["lancement15minutes"] == 15:
            await update_sheet("15m", quinsesminutesdebut, 19, save_csv)
            quinsesminutesdebut += 1
            lancements["lancement15minutes"] = 0

        if lancements["lancement30minutes"] == 30:
            await update_sheet("30m", trenteminutesdebut, 25, save_csv)
            trenteminutesdebut += 1
            lancements["lancement30minutes"] = 0

        if lancements["lancement1h"] == 60:
            await update_sheet("1h", uneheuredebut, 31, save_csv)
            uneheuredebut += 1
            lancements["lancement1h"] = 0

        if lancements["lancement2h"] == 120:
            await update_sheet("2h", deuxheuresdebut, 37, save_csv)
            deuxheuresdebut += 1
            lancements["lancement2h"] = 0

        if lancements["lancement4h"] == 240:
            await update_sheet("4h", quatreheuresdebut, 43, save_csv)
            quatreheuresdebut += 1
            lancements["lancement4h"] = 0

        if lancements["lancement6h"] == 360:
            await update_sheet("6h", sixheuresdebut, 49, save_csv)
            sixheuresdebut += 1
            lancements["lancement6h"] = 0

        if lancements["lancement8h"] == 480:
            await update_sheet("8h", huitheuresdebut, 55, save_csv)
            huitheuresdebut += 1
            lancements["lancement8h"] = 0

        if lancements["lancement12h"] == 720:
            await update_sheet("12h", douzeheuresdebut, 61, save_csv)
            douzeheuresdebut += 1
            lancements["lancement12h"] = 0

        if lancements["lancement1jour"] == 1440:
            await update_sheet("1d", unedjoudebut, 67, save_csv)
            unedjoudebut += 1
            lancements["lancement1jour"] = 0

        if lancements["lancement3jour"] == 4320:
            await update_sheet("3d", troisjousdebut, 73, save_csv)
            troisjousdebut += 1
            lancements["lancement3jour"] = 0

        if lancements["lancement1semaine"] == 10080:
            await update_sheet("1w", une_semaine_debut, 79, save_csv)
            une_semaine_debut += 1
            lancements["lancement1semaine"] = 0

        await asyncio.sleep(60)

def print_banner():
    banner = (
        f"{YELLOW}    ____ ______________________  ___   ________ __ __________ {NORMAL}\n"
        f"{YELLOW}   / __ )_  __/ ____/_  __/ __ \\/   | / ____/ //_// ____/ __ \\{NORMAL}\n"
        f"{YELLOW}  / __  |/ / / /     / / / /_/ / /| |/ /   / ,<  / __/ / /_/ /{NORMAL}\n"
        f"{YELLOW} / /_/ // / / /___  / / / _, _/ ___ / /___/ /| |/ /___/ _, _/ {NORMAL}\n"
        f"{YELLOW}/_____//_/  \\____/ /_/ /_/ |_/_/  |_\\____/_/ |_/_____/_/ |_|  {NORMAL}\n"
        f"                                            {YELLOW}B T C T R A C K E R{NORMAL}\n"
    )
    print(banner)

def print_output_options():
    print(f"{YELLOW}Choisissez le format de sortie :{NORMAL}")
    print(f"{NORMAL}[{RED}1{NORMAL}]  CSV")
    print(f"{NORMAL}[{RED}2{NORMAL}]  Excel")
    print(f"{NORMAL}[{RED}3{NORMAL}]  Les deux")

def choose_format():
    print_banner()
    print_output_options()
    choice = input("Entrez votre choix (1/2/3) : ")

    save_csv = False
    save_excel = False

    if choice == '1':
        print("Les données seront uniquement enregistrées en format CSV.")
        save_csv = True
    elif choice == '2':
        print("Les données seront uniquement enregistrées en format Excel.")
        save_excel = True
    elif choice == '3':
        print("Les données seront enregistrées au format CSV et Excel.")
        save_csv = True
        save_excel = True
    else:
        print("Choix invalide. Les données seront enregistrées au format CSV par défaut.")
        save_csv = True

    return save_csv, save_excel

# Configuration de la boucle d'événements pour Windows
if __name__ == "__main__":
    import platform

    if platform.system() == 'Windows':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    save_csv, save_excel = choose_format()
    if save_excel:
        save_to_excel()
    asyncio.run(main(save_csv, save_excel))
